import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import numpy as np

# DataFrames
df_produtos = pd.DataFrame({
    'produto': ['Celular', 'Notebook', 'Monitor', 'Teclado'],
    'preço': [2500, 4000, 800, 150]
})

nomes_aleatorios = ['Carlos', 'Mariana', 'Ricardo', 'Julia', 'Felipe']
df_usuarios = pd.DataFrame({
    'Nome': np.random.choice(nomes_aleatorios, size=3),
    'Idade': np.random.randint(20, 50, size=3)
})

# --- PASSO 1: Salvar APENAS OS DADOS dos produtos na linha 3, coluna B ---

writer = pd.ExcelWriter('dados_formatados_final.xlsx', engine='openpyxl')

# startcol=1 (Coluna B), startrow=2 (Linha 3), header=False (sem cabeçalho)
df_produtos.to_excel(writer, index=False, startcol=1, startrow=2, header=False)

writer.close()

# --- PASSO 2: Reabrir e adicionar o cabeçalho na linha 2 ---

workbook = openpyxl.load_workbook('dados_formatados_final.xlsx')
worksheet = workbook.active
fonte_cabecalho = Font(bold=True, color='0000FF')

# Adiciona os cabeçalhos do DataFrame de produtos na linha 2
# O loop começa na coluna 2 (índice 2) para garantir o início na coluna B
for col_num, coluna in enumerate(df_produtos.columns, 2):
    cell = worksheet.cell(row=2, column=col_num, value=coluna)
    cell.font = fonte_cabecalho

# --- PASSO 3: Adicionar o DataFrame de usuários com a mesma lógica ---

# Encontra a próxima linha vazia (depois dos dados do 1º DataFrame) e adiciona 2 para o espaçamento
proxima_linha = worksheet.max_row + 2

# Adiciona o título "Usuários"
# O título também inicia na coluna B (coluna 2)
titulo_cell = worksheet.cell(row=proxima_linha, column=2, value="Usuários")
titulo_cell.font = Font(bold=True)
titulo_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Adiciona os cabeçalhos do novo DataFrame, começando da coluna B (índice 2)
linha_dados_usuarios = proxima_linha + 1
for col_num, coluna in enumerate(df_usuarios.columns, 2):
    cell = worksheet.cell(row=linha_dados_usuarios, column=col_num, value=coluna)
    cell.font = Font(bold=True, color='0000FF')

# Adiciona os dados do novo DataFrame, também a partir da coluna B (índice 2)
for row_num, row_data in enumerate(df_usuarios.values, linha_dados_usuarios + 1):
    for col_num, valor in enumerate(row_data, 2):
        worksheet.cell(row=row_num, column=col_num, value=valor)

# Salva o arquivo final
workbook.save('dados_formatados_final.xlsx')

print("Arquivo Excel atualizado! Verifique o arquivo 'dados_formatados_final.xlsx'.")